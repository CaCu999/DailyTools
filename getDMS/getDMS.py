import tkinter as tk
from tkinter import filedialog
import xlrd.sheet
import pandas as pd
import os
import json
import threading
from openpyxl import load_workbook
import xlrd
from typing import Tuple
from datetime import datetime

window = tk.Tk()
dmsVar = tk.StringVar()
msgVar = tk.StringVar()
dataVar = tk.StringVar()
excelVar = tk.StringVar()
BAVar = tk.StringVar()
out_text = tk.Text(window)


def selectDMSPath():
    path  = filedialog.askdirectory(initialdir=history["dms"])
    if path == "":
        return
    dmsVar.set(path)
    history["dms"] = path

def selectExcelPath():
    file  = filedialog.askopenfilename(initialdir= os.path.dirname(history["excel"]))
    if file == "":
        return
    excelVar.set(file)
    history["excel"] = file

def selectBAPath():
    file  = filedialog.askopenfilename(initialdir= os.path.dirname(history["BA"]))
    if file != "":
        BAVar.set(file)
        history["BA"] = file
        threading.Thread(target=readBA, args={file}).start()

def log_pnt(input):
    input = str(input)
    out_text.insert('end', input + "\n")
    out_text.yview('end')

def readBA(file):
    global ba,ba_flag
    ba_flag = True
    ba = pd.read_excel(io=file, engine="openpyxl", sheet_name="ビットアサイン表", header=10, skiprows=[10])
    out_text.insert("end", "read succeed!\n")
    ba_flag = False
    # print(ba)

def readDMSExcel(msg, data, file):
    log_pnt(file)
    workbook = xlrd.open_workbook_xls(filename=file)
    sh = workbook.sheet_by_index(0)
    log_pnt(f'\n\n==========={msg}.{data}============')
    log_pnt(f'-------------english-------------')
    log_pnt(sh.cell_value(35, 2))
    log_pnt(f'-------------japanese-------------')
    log_pnt(sh.cell_value(38, 2))

def readDMSExcel(msg, data, file)->Tuple[str, str]:
    log_pnt(file)
    workbook = xlrd.open_workbook_xls(filename=file)
    sh = workbook.sheet_by_index(0)
    eng = sh.cell_value(35, 2)
    jp = sh.cell_value(38, 2)
    log_pnt(f'\n\n==========={msg}.{data}============')
    log_pnt(f'-------------english-------------')
    log_pnt(eng)
    log_pnt(f'-------------japanese-------------')
    log_pnt(jp)
    return eng, jp

def selectFromVar():
    msg = msgVar.get()
    data = dataVar.get()
    if msg == "" or data == "":
        log_pnt("null message or null label")
        return
    dms = dmsVar.get()
    if dms == "":
        log_pnt("null dms path")
        return
    if ba_flag:
        log_pnt("waiting for read BA excel")
        return
    #save history
    js = json.dumps(history)
    with open(hisfile, "w", encoding="utf-8") as f:
        f.write(js)
    local_file = checkFileName(msg, data, dms)
    res = os.path.exists(local_file)
    if res:
        threading.Thread(target=readDMSExcel, args={msg, data, loca_file}).start()
    else:
        log_pnt(f'{dms}: {res}')

def checkFileName(msg, data, filename) -> str:
    if ba_flag:
        return ""
    dms_str = str(ba.loc[ba['Data Label'] == data, 'Data\nMaster\nSheet'].values[0])
    # log_pnt(ba.loc[ba['Data Label'] == data, 'Data\nMaster\nSheet'].values[0])
    loca_file = os.path.join(filename, msg, f'{data}-{dms_str}.xls')
    # print(ba['Data\nMaster\nSheet'][])
    return loca_file

def selectFromFile():
    if ba_flag:
        log_pnt("reading BA xlsx ing......\npress after reading succeed.")
        return
    file = excelVar.get()
    dmsFile = dmsVar.get()
    book = load_workbook(filename=file)
    sh = book['Sheet1']
    log_pnt(sh.columns)
    for r in range(2, sh.max_row + 1):
        msg = sh.cell(row=r, column=3).value
        data = sh.cell(row=r, column=4).value
        if msg == None or data == None:
            continue
        log_pnt(f'{msg}.{data}')
        path = checkFileName(msg, data, dmsFile)
        [eng, jp] = readDMSExcel(msg, data, path)
        celleng = sh.cell(row=r, column=6)
        celljp = sh.cell(row=r, column=7)
        celleng.value = eng
        celljp.value = jp
    dir = os.path.dirname(file)
    name, extension = os.path.splitext(os.path.basename(file))

    log_pnt(f'{dir}-{name}  {extension}')
    cur_time = datetime.now().strftime("%H%M")
    new_file = os.path.join(dir, f'{name}_{cur_time}{extension}')
    new_file.replace('\\','/')
    book.save(new_file)
    log_pnt(f'file generated: {os.path.basename(new_file)}')

def getJson():
    global hisfile, history
    hisfile = os.path.join(os.path.dirname(__file__), "file", "history.json")
    with open(hisfile) as f:
        history = json.load(f)
    dmsVar.set(history["dms"])
    excelVar.set(history["excel"])
    BAVar.set(history["BA"])

def mainWindow():
    window.geometry("500x400")
    getJson()
    if history['BA'] == "":
        file = os.path.join(os.path.dirname(__file__), "file", "BA.xlsx")
    else:
        file = history['BA']
    threading.Thread(target=readBA, args={file}).start() 
    line_height = 0.08
    # line 1
    line = 0
    dmsLabel = tk.Label(window, text="DMS路径")
    dmsLabel.place(relheight=line_height, relwidth=0.1, relx=0.01, rely=0.03 + line * 0.1)
    dmsPathText = tk.Entry(window, textvariable=dmsVar)
    dmsPathText.place(relheight=line_height, relwidth=0.7, relx=0.13, rely=0.03 + line * 0.1)
    dmsbutton = tk.Button(window,text="选择", command=selectDMSPath)
    dmsbutton.place(relheight=line_height, relwidth=0.1, relx=0.85, rely=0.03 + line * 0.1)
    
    # line 2
    line = 1
    excelPath = tk.Label(window, text="文档路径")
    excelPath.place(relheight=line_height, relwidth=0.1, relx=0.01, rely=0.03 + line * 0.1)
    excelPathText = tk.Entry(window, textvariable=excelVar)
    excelPathText.place(relheight=line_height, relwidth=0.7, relx=0.13, rely=0.03 + line * 0.1)
    excelPathbutton = tk.Button(window,text="选择", command=selectExcelPath)
    excelPathbutton.place(relheight=line_height, relwidth=0.1, relx=0.85, rely=0.03 + line * 0.1)
    
    # line 3
    line = 2
    BAPath = tk.Label(window, text="BA路径")
    BAPath.place(relheight=line_height, relwidth=0.1, relx=0.01, rely=0.03 + line * 0.1)
    BAPathText = tk.Entry(window, textvariable=BAVar)
    BAPathText.place(relheight=line_height, relwidth=0.7, relx=0.13, rely=0.03 + line * 0.1)
    BAPathbutton = tk.Button(window,text="选择", command=selectBAPath)
    BAPathbutton.place(relheight=line_height, relwidth=0.1, relx=0.85, rely=0.03 + line * 0.1)

    # line 4
    line = 3
    msgLabel = tk.Label(window, text="Message")
    msgLabel.place(relheight=line_height, relwidth=0.13, relx=0.01, rely=0.03 + line * 0.1)
    msgEntry = tk.Entry(window, textvariable=msgVar)
    msgEntry.place(relheight=line_height, relwidth=0.25, relx=0.15, rely=0.03 + line * 0.1)
    dataLabel = tk.Label(window, text="Data")
    dataLabel.place(relheight=line_height, relwidth=0.13, relx=0.51, rely=0.03 + line * 0.1)
    dataEntry = tk.Entry(window, textvariable=dataVar)
    dataEntry.place(relheight=line_height, relwidth=0.25, relx=0.65, rely=0.03 + line * 0.1)

    # line 5
    line = 4
    strictVal = tk.Button(window,text="从输入框读取", command=selectFromVar)
    strictVal.place(relheight=line_height, relwidth=0.35, relx=0.05, rely=0.03 + line * 0.1)
    fileVal = tk.Button(window,text="读取文件生成", command=selectFromFile)
    fileVal.place(relheight=line_height, relwidth=0.35, relx=0.55, rely=0.03 + line * 0.1)

    # line 6
    line = 5
    out_text.place(relheight=0.4, relwidth=0.9, relx=0.05, rely=0.05 + line * 0.1)



    window.mainloop()
    # logtag.place(relwidth=0.73, relheight=0.3, relx=0.4, rely=0)

if __name__ == "__main__":
    mainWindow()