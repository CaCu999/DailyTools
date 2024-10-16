import pandas as pd
import xlwings as xw
import tkinter as tk
from tkinter import ttk
import queue
import time
import os
from tkinter import filedialog
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font,Border,Side
import shutil
import threading
import math

#config -------
dirList = ["Project Root Directory", "Product Name", "Case Predix", "Screens Directory"]
funcList = [ "Project Root Directory","Checklist Path","Product Name",
    "Case Predix", "Screen Table","Screens Directory", "Scripts Table", 
    "Can Config", "Pic Mapping Table", "Api Mapping Table"]
xlsxMap = {}
#adb python
abs = 'def getProjectPath() :\n'
abs += '\treturn __file__[:__file__.find("24mm") + len("24mm")]\n\n'
abs += 'import sys\n'
abs += 'if not (getProjectPath() in sys.path) :\n'
abs += '\tsys.path.append(getProjectPath())\n'
abs += 'print("init " + __file__)\n'
#case python header
case_header = 'from autost.api import *\n'
case_header += 'import AbsPath\n\n'
case_header += 'from Common.InfraLibAPI import *'
root_path = '../../../../../../../../AutoTestCase/24mm/'
# root_path = 'D:/Project/PythonProject/pac'
#uilist -------
#功能label
settingLabels = {}
#显示值label
resultLabels = {}
resultVars = {}
#设置值edit
editEntities = {}
editVars = {}
#发送结果
buttonItems = {}
#清空结果
buttonResetItems = {}
window = tk.Tk()
log_text = tk.Text(window)
info_text = tk.Text(window)
scrollarbar = tk.Scrollbar(window)
q = queue.Queue()
target = ""
#cell type
yellow_fill = PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
bold_font = Font(bold=True, size=11)
all_border = Border(left=Side(style="thin", color='000000'), right=Side(style="thin", color='000000'),
                bottom=Side(style="thin", color='000000'), top=Side(style="thin", color='000000'))
#file select -----
def dirSelect(item):
    dir = filedialog.askdirectory()
    editVars[item].set(dir)

def fileSelect(item):
    file = filedialog.askopenfile()
    editVars[item].set(file.name)

#ui setting
def set_textVar(textShow, isClear = False):
    # with open(logpath, 'a') as f:
    textShow.config(state="normal")
    if isClear:
        textShow.delete('1.0','end')
    while not q.empty():
        str = q.get()
        # f.write(str)
        textShow.insert('end', str)
        textShow.yview('end')
    textShow.config(state='disabled')

def logpnt(log):
    tr = str(log) + "\n"
    q.put(tr)

def update_ui():
    set_textVar(log_text)
    window.after(50, update_ui)
    
def mainWindow():
    #prepare excel and read DataFrame
    global path,logdir,logpath,filename,sh
    #set window
    window.geometry("1000x900")
    style = ttk.Style(window)
    style.configure('lefttab.TNotebook', tabposition='wn')
    panel = tk.Frame(window)
    # notebook.grid(row=0, column=0)
    panel.place(relx=0, rely=0, relwidth=1, relheight=0.5)
    panel.columnconfigure(0, weight=3)
    panel.columnconfigure(1, weight=20)
    panel.columnconfigure(2, weight=1)
    panel.columnconfigure(3, weight=2)
    log_text.place(relx=0, rely=0.5, relwidth=0.98, relheight=0.5)
    log_text.config(state="disabled")
    scrollarbar.place(relx=0.98, rely=0.5, relheight=0.5)
    log_text.config(yscrollcommand=scrollarbar.set)
    scrollarbar.config(command=log_text.yview)
    for i,item in enumerate(funcList):
        settinglabel = tk.Label(panel, text=item + ': ')
        editCms = tk.StringVar(value=str(json_data[item]))
        editVars[item] = editCms
        editEntity = tk.Entry(panel, textvariable=editCms)
        # editEntity.bind("<Button-1>", func=lambda event, item=item:editClick(item))
        editEntity.focus_force()
        if item in dirList:
            button = tk.Button(panel, text="...", command=lambda item=item: dirSelect(item))
        else:            
            button = tk.Button(panel, text="...", command=lambda item=item: fileSelect(item))
        settingLabels[item] = settinglabel
        editEntities[item] = editEntity
        buttonItems[item] = button
        settinglabel.grid(row=i, column=0, sticky='ew')
        editEntity.grid(row=i, column=1, sticky='ew')
        button.grid(row=i, column=3, sticky='ew')
    button_start = tk.Button(panel, text="generate", command=start ) 
    button_start.grid(row=len(funcList) + 1, column=3, sticky='ew')
    window.after(50, update_ui)
    window.mainloop()

# func region --------
def threading_start(wb, file, target):
    ws = wb.active
    ng_idx = checklist.columns.size

    condition_idx, action_idx, check_idx, case_idx, auto_idx = get_condition_index(ws)
    print(f'con:{condition_idx}, act:{action_idx}, check:{check_idx}')
    i = 0
    for row in ws.iter_rows(values_only=False , min_row=2):
        # print(row)
        i += 1
        per = f'[{i}/{checklist.shape[0]}] '
        global row_reason
        row_reason = ''
        case_content = ''
        condition = row[condition_idx].value
        action = row[action_idx].value
        check = row[check_idx].value
        case_name = row[case_idx].value
        if row[condition_idx].font.strike or row[auto_idx].value == '否':
            insert_result(row, 'passed', True, ng_idx)
            logpnt(f'{per}{case_name}:  passed~~~\n\n')
            continue
        #generate case py
        case_content = case_header
        case_content += filter_condition(condition, case_name, row, ng_idx)
        case_content += filter_action(action, case_name, row, ng_idx)
        case_content += filter_check(check, case_name, row, ng_idx)
        row_reason = row_reason.strip()
        if row_reason == '':
            #generate case py
            generate_case_py(file, case_name, case_content)
            #generate config ini
            #generate AbsPath py
            generate_Abs_py(file, case_name)
            insert_result(row, case_name + '.tcs', True, ng_idx)
            logpnt(f'{per}{case_name} : succeed!!!\n\n')
        else:
            insert_result(row, row_reason, False, ng_idx)
            logpnt(f'{per}{case_name} : failed')
            logpnt(row_reason + '\n\n')
            continue
    #print(case_content)
    wb.save(target)
    predix_name = json_data["Case Predix"]
    tar_dir = os.path.join(os.path.dirname(file), 'AutoCase', predix_name)
    if not os.path.exists(tar_dir):
        os.makedirs(tar_dir)
    new_path = os.path.join(tar_dir, os.path.basename(file))
    wb.save(new_path)
    logpath = os.path.join(tar_dir, f'case_generate_{now}.log')
    with open(logpath, 'a') as f:
        str = log_text.get('1.0', 'end')
        f.write(str)
    logpnt('---------------succeed and saved.--------------')


# start generate
def start():
    #save json
    for item in funcList:
        json_data[item] = editVars[item].get()
    with open('testPanel/config.json', 'w', encoding='utf-8')  as f:
        json.dump(json_data, f, ensure_ascii=False, indent=4)
    log_text.config(state="normal")
    log_text.delete('1.0','end')
    log_text.config(state="disabled")
    global checklist,screeninfo,vehicletype,picmapping, script
    #read pandas
    checklist = pd.read_excel(io=json_data["Checklist Path"], sheet_name="测试用例", engine="openpyxl")
    screeninfo = pd.read_excel(io=json_data['Screen Table'], sheet_name='Sheet1', engine='openpyxl')
    picmapping = pd.read_excel(io=json_data['Pic Mapping Table'], sheet_name='PicMapping', engine='openpyxl')
    vehicletype = json_data["VehicleType"]
    script = pd.read_excel(io=json_data["Scripts Table"], sheet_name='Utilities', engine='openpyxl')
    # copy file
    file, target = get_result_name()
    wb,ws = add_title(file)
    reason_idx = checklist.columns.size + 1
    case_path = os.path.join(os.path.dirname(file), 'AutoCase', json_data["Case Predix"])
    shutil.rmtree(case_path) if os.path.exists(case_path) else print('not exist')
    os.mkdir(case_path)
    t = threading.Thread(target=threading_start, args=(wb,file, target))
    t.daemon = True
    t.start()
    

# get target file name
def get_result_name():
    file = json_data["Checklist Path"]
    target = f'{os.path.splitext(file)[0]}_{now}{os.path.splitext(file)[1]}'
    return file, target

# add result columns
def add_title(file):
    title_num = checklist.columns.size
    wb = load_workbook(file)
    ws = wb.active
    new_col = ['生成脚本','生成脚本NG原因']
    for i, col in enumerate(new_col, start=1):
        ws.insert_cols(i + title_num)
        cell = ws.cell(row=1, column = i + title_num)
        cell.value = col
        cell.font = bold_font
        cell.fill = yellow_fill
        cell.border = all_border 
    return wb,ws

def insert_result(ws, item, res, idx):
    # print(ws[idx].value)
    cell_ng = ws[idx]
    cell_reason = ws[idx + 1]
    cell_ng.border = all_border
    cell_reason.border = all_border
    cell_ng.value = 'OKKKKKKKK'
    if res:
        if item == 'passed':
            cell_ng.value = item
            cell_reason.value = "case is deleted"
        else:
            cell_ng.value = item
    else:
        cell_ng.value = 'NG'
        cell_reason.value = item
        # cell_reason.value += cell_reason.value + title + item + '\n'
    return '\n'


def get_condition_index(ws):
    con = 0
    action = 0
    check = 0
    case = 0
    auto = 0
    for row in ws.iter_rows(values_only=True, max_row=1):
        for item in enumerate(row):
            title = str(item[1])
            if title.lower() == 'condition':
                con = item[0]
            elif title.lower() == 'action':
                action = item[0]
            elif title.lower() == 'check':
                check = item[0]
            elif title.lower() == 'autocaseid':
                case = item[0]
            elif title == '是否可自动化':
                auto = item[0]
    return con, action, check, case, auto

def generate_case_py(file, case, content):
    dir = os.path.dirname(file)
    predix_name = json_data["Case Predix"]
    path = os.path.join(dir, 'AutoCase', predix_name, case + '.tcs')
    if not os.path.exists(path):
        os.makedirs(path)
    path = os.path.join(path, case + '.py')
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)

def generate_Abs_py(file, case):
    dir = os.path.dirname(file)
    predix_name = json_data["Case Predix"]
    path = os.path.join(dir, 'AutoCase', predix_name, case+'.tcs')
    if not os.path.exists(path):
        os.makedirs(path)
    path = os.path.join(path, 'AbsPath.py')
    with open(path, 'w') as f:
        f.write(abs)

def filter_condition(condition, case, row, ng):
    content = "\n#前置条件\n"
    conds = condition.split("\n")
    global row_reason
    for i,item in enumerate(conds):
        content += f'#{item}\n'
        item = item[2:]
        item = item.strip()
        if item.startswith("发送CAN信号"):
            res = can_2_api(item)
        elif item.startswith("诊断读取CAN信号"):
            res = can_2_api(item, True)
        elif item.startswith('进入') or item.startswith('在') or item.startswith('确认'):
            res = img_2_api(item)
        elif item.startswith('等待'):
            res = wait_2_api(item)
        else:
            res = script_2_api(item)
        content += read_res(res, '[前置条件]' + item, '未找到脚本')
    return content

def filter_action(condition, case, row, ng):
    content = "\n\n\n#手顺\n"
    conds = condition.split("\n")
    for i,item in enumerate(conds):
        content += f'#{item}\n'
        item = item[2:]
        item = item.strip()
        res = None
        if item.startswith("发送CAN信号"):
            res = can_2_api(item)
        elif item.startswith("诊断读取CAN信号"):
            res = can_2_api(item, True)
        elif item.startswith("语音输入"):
            res = voice_2_api(item)
        elif item.startswith('进入') or item.startswith('在') or item.startswith('确认'):
            res = img_2_api(item)
        elif item.startswith('等待'):
            res = wait_2_api(item)
        else:
            res = script_2_api(item)
        content += read_res(res, '[手顺]' + item, '未找到脚本')
    return content

def filter_check(condition, case, row, ng):
    content = "\n\n\n#期待结果\n"
    conds = condition.split("\n")
    for i,item in enumerate(conds):
        content += f'#{item}\n'
        item = item[2:]
        item = item.strip()
        res = None
        scr_res = read_Value(script, '别名', item)
        if scr_res != -1:
            res = script_2_api(item)
        elif item.startswith("发送CAN信号"):
            res = can_2_api(item)
        elif item.startswith("诊断读取CAN信号"):
            res = can_2_api(item, True)
        elif item.startswith('进入') or item.startswith('在') or item.startswith('确认'):
            res = img_2_api(item)
        elif item.startswith('等待'):
            res = wait_2_api(item)
        content += read_res(res, '[手顺]' + item, '未找到图片')
    return content

def read_Value(df, title, value):
    res = df[df[title] == value]
    if res.empty:
        logpnt(f'read {value} in {title} --> {-1}')
        return -1
    logpnt(f'read {value} in {title} --> {res.index[0]}')
    return res.index[0]

def read_res(res, item, reason):
    global row_reason
    if res is not None:
        content = res + '\n'
    else:
        row_reason += f'{item}({reason})\n'
        content = '\n'
    return content

def read_img_idx(img):
    idx = read_Value(screeninfo, '别名', img)
    if idx == -1:
        idx = read_Value(screeninfo, '画面名', img)
    return idx

def read_img_state(path, screen, img):
    screen_dir = json_data["Screen Table"]
    path = os.path.join(path, screen, 'elements', 'pics', vehicletype)
    img = img.replace('是显示状态', '')
    img = img.replace('是非显示状态', '')
    if img.__contains__('是'):
        state = img.split('是')
        img = f'{state[0]}_{state[1]}'
    elif img.endswith('显示'):
        img = img[:len(img)-2]
    elif img.__contains__('的'):
        state = img.split('的')
        img = f'{state[1]}_{state[0]}'
    res = read_Value(picmapping, '用例图片名', img)
    if res == -1:
        return None
    img = picmapping['素材图片名'][res]
    state = img + '.png'
    path = os.path.join(path, state)
    path = path.replace('\\', '/')
    return path

def can_2_api(condition, isRead = False):
    # assertReadCAN("AVN1S05.S_TPCD_D.phys", 0)
    config = pd.read_excel(io=json_data["Can Config"], sheet_name='CAN Config', engine="openpyxl")
    result = None
    idx = read_Value(config, "case_description", condition)
    if not isRead:
        if idx != -1:
            m = config['name'][idx]
            v = config['value'][idx]
            i = config['interval'][idx]
            ch = config['channel'][idx]
            try:
                r = float(ch)
            except ValueError:
                r = None
            if r is not None and math.isnan(r):
                result = f'sendCAN(\'{m}\', {v}, interval={i})\n'
            else:
                result = f'sendCAN(\'{m}\', {v}, interval={i}, channel=\'{ch}\')\n'
        else:
            condition = condition.replace(' ', '')
            condition = condition.replace('：',":")
            condition = condition.replace('（',"(")
            condition = condition.split(":")[1]
            if condition.__contains__('('):
                condition = condition.split('(')[0]
            if not condition.__contains__('=') or not condition.__contains__('.'):
                return None
            msg = condition.split("=")[0] + '.phys'
            value = condition.split("=")[1]
            value = value.lower()
            if value.endswith('b'):
                value = value[0:len(value)-1]
                value = int(value, 2)
            elif value.endswith('h'):
                value = value[0:len(value)-1]
                value = int(value, 16)
            elif value.endswith('d'):
                value = value[0:len(value)-1]
                value = int(value)
            else:
                value = int(value)
            result = f'sendCAN(\'{msg}\', {value}, interval={1})\n'
    else:
        if idx != -1:
            m = config['name'][idx]
            v = config['value'][idx]
            i = config['interval'][idx]
            ch = config['channel'][idx]
            try:
                r = float(ch)
            except ValueError:
                r = None
            if r is not None and math.isnan(r):
                result = f'assertReadCAN(\'{m}\', {v})\n'
            else:
                result = f'assertReadCAN(\'{m}\', {v}, channel=\'{ch}\')\n'
        else:
            condition = condition.replace(' ', '')
            condition = condition.replace('：',":")
            condition = condition.replace('（',"(")
            condition = condition.split(":")[1]
            if condition.__contains__('('):
                condition = condition.split('(')[0]
            if not condition.__contains__('=') or not condition.__contains__('.'):
                return None
            msg = condition.split("=")[0] + '.phys'
            value = condition.split("=")[1]
            value = value.lower()
            if value.endswith('b'):
                value = value[0:len(value)-1]
                value = int(value, 2)
            elif value.endswith('h'):
                value = value[0:len(value)-1]
                value = int(value, 16)
            elif value.endswith('d'):
                value = value[0:len(value)-1]
                value = int(value)
            else:
                value = int(value)
            result = f'assertReadCAN(\'{msg}\', {value})\n'
    return result

def voice_2_api(condition):
    condition = condition.replace(' ', '')
    condition = condition.replace('：',":")
    condition = condition.replace('（',"(")
    condition = condition.split(":")[1] 
    result = None
    root = json_data['Project Root Directory']
    voicepath = os.path.join(root, 'Common/InfraLib/VRMaterials/audios', condition + '.wav')
    voicepath = voicepath.replace('\\','/')
    if os.path.exists(voicepath):
        logpnt(os.path.basename(voicepath) + '存在')
        voicepath = voicepath.replace(root, root_path)
        result = f'say(Audio("{voicepath}"))'
    return result

def script_2_api(condition):
    idx = read_Value(script, '别名', condition)
    if idx == -1:
        return None
    file = str(script['脚本路径'][idx])
    if file == "":
        return None
    file = file.replace('\\', '/')
    result = f'do_segment(Segment("../../../../../../../../AutoTestCase/24mm/{file}"))'
    return None if result is None else result + '\n'

def img_2_api(condition):
    root = json_data['Project Root Directory']
    screen = json_data['Screens Directory'].split(root)[1]
    path = root + screen  
    path = path.replace('/AirConditioner', '')  
    result = None
    if condition.startswith('进入'):
        img = str(condition.split('进入')[1])
        idx = read_img_idx(img)
        if idx == -1:
            return None
        path = os.path.join(path, screeninfo['OPE'][idx], screeninfo['画面ID'][idx], 'flows','flow.tcs')
        path = path.replace('\\', '/')
        if os.path.exists(path):
            path = path.replace(root, root_path)
            result = f'do_segment(Segment("{path}"))'
    elif condition.startswith('确认'):
        img = condition.split('确认')[1]
        if img.endswith('显示'):
            img = img.replace('显示', '')
        idx = read_img_idx(img)
        if idx == -1:
            return None
        # TBD 画面位置
        path = os.path.join(path, screeninfo['OPE'][idx], screeninfo['画面ID'][idx])
        path = path.replace('\\','/')
        path = path.replace(root, root_path)
        result = f'assert_exists(Template({path}))'
    elif condition.startswith('在') and condition.__contains__('点击'):
        idx = read_img_idx(condition[1:].split('上')[0])
        if idx == -1:
            return None
        if not os.path.exists(os.path.join(path, screeninfo['OPE'][idx], screeninfo['画面ID'][idx])):
            return None
        path = os.path.join(path, screeninfo['OPE'][idx])
        img = read_img_state(path, screeninfo['画面ID'][idx], condition.split('点击')[1])
        if img is None  or not os.path.exists(img):
            return None        
        img = img.replace(root, root_path)
        result = f'touch_if(Template("{img}"))'
    elif condition.startswith('在') and condition.endswith('是非显示状态'):
        idx = read_img_idx(condition[1:].split('上')[0])
        if idx == -1:
            return None
        if not os.path.exists(os.path.join(path, screeninfo['OPE'][idx], screeninfo['画面ID'][idx])):
            return None
        path = os.path.join(path, screeninfo['OPE'][idx])
        img = read_img_state(path, screeninfo['画面ID'][idx], condition.split('确认')[1])
        if img is None or not os.path.exists(img):
            return None        
        img = img.replace(root, root_path)
        result = f'assert_not_exists(Template("{img}"))'
    elif condition.startswith('在') and condition.__contains__('确认'):
        idx = read_img_idx(condition[1:].split('上')[0])
        if idx == -1:
            return None
        if not os.path.exists(os.path.join(path, screeninfo['OPE'][idx], screeninfo['画面ID'][idx])):
            return None
        path = os.path.join(path, screeninfo['OPE'][idx])
        img = read_img_state(path, screeninfo['画面ID'][idx], condition.split('确认')[1])
        if img is None  or not os.path.exists(img):
            return None        
        img = img.replace(root, root_path)
        result = f'assert_exists(Template("{img}"))'        
    return result

def wait_2_api(condition):
    result = None
    condition = condition.strip()
    logpnt(condition)
    condition = condition.replace('等待', '')
    if condition.endswith('ms'):
        t = int(condition.replace('ms', ''))
        t = float(t * 0.001)
        result = f'sleep({t})'
    elif condition.endswith('s'):
        t = int(condition.replace('s', ''))
        result = f'sleep({t})'
    return result

if __name__ == "__main__":
    global json_data,now
    now = str(time.strftime("%d%H%M%S", time.localtime()))
    print(now)
    with open("testPanel/config.json", encoding='utf-8')  as f:
        json_data = json.load(f)
    
    mainWindow()
    # start()